from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from models import DatasetFieldType, _coerce_dataset_value


@dataclass
class ParsedDataset:
    field_names: list[str]
    rows: list[list[Any]]
    inferred_types: dict[str, DatasetFieldType]


def parse_dataset_file(path: str) -> ParsedDataset:
    suffix = Path(path).suffix.lower()
    if suffix == '.csv':
        field_names, rows = _parse_csv(path)
    elif suffix in {'.xlsx', '.xlsm'}:
        field_names, rows = _parse_xlsx(path)
    else:
        raise ValueError('Unsupported file type. Use .csv or .xlsx.')

    if not field_names:
        raise ValueError('No header row found.')
    _validate_header(field_names)
    inferred = {name: _infer_type_for_column(rows, index) for index, name in enumerate(field_names)}
    return ParsedDataset(field_names=field_names, rows=rows, inferred_types=inferred)


def validate_and_convert_rows(
    parsed: ParsedDataset,
    selected_types: dict[str, DatasetFieldType],
) -> list[dict[str, Any]]:
    if set(selected_types.keys()) != set(parsed.field_names):
        raise ValueError('Selected field types do not match inferred fields.')

    converted_rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(parsed.rows, start=2):
        if len(row) != len(parsed.field_names):
            raise ValueError(f'Row {row_index} has wrong number of columns.')
        converted_row: dict[str, Any] = {}
        for col_index, field_name in enumerate(parsed.field_names):
            field_type = selected_types[field_name]
            raw_value = row[col_index]
            if _is_empty(raw_value):
                raise ValueError(f'Row {row_index}, field "{field_name}" is empty. Nulls are not allowed.')
            value_int, value_float, value_bool, value_string = _coerce_dataset_value(field_type, raw_value)
            if field_type == DatasetFieldType._int:
                converted_row[field_name] = value_int
            elif field_type == DatasetFieldType._float:
                converted_row[field_name] = value_float
            elif field_type == DatasetFieldType._bool:
                converted_row[field_name] = value_bool
            else:
                converted_row[field_name] = value_string
        converted_rows.append(converted_row)
    return converted_rows


def _parse_csv(path: str) -> tuple[list[str], list[list[Any]]]:
    with open(path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return [], []
    header = [str(cell).strip() for cell in rows[0]]
    data_rows = [list(row) for row in rows[1:] if any(str(cell).strip() != '' for cell in row)]
    return header, data_rows


def _parse_xlsx(path: str) -> tuple[list[str], list[list[Any]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise ValueError('XLSX import requires openpyxl installed.') from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return [], []
    header = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
    data_rows = []
    for row in rows[1:]:
        row_values = list(row)
        if any(not _is_empty(cell) for cell in row_values):
            data_rows.append(row_values)
    return header, data_rows


def _validate_header(field_names: list[str]) -> None:
    cleaned = [name.strip() for name in field_names]
    if any(name == '' for name in cleaned):
        raise ValueError('Header contains empty field name.')
    if len(set(cleaned)) != len(cleaned):
        raise ValueError('Header contains duplicate field names.')


def _infer_type_for_column(rows: list[list[Any]], index: int) -> DatasetFieldType:
    values = [row[index] for row in rows if index < len(row) and not _is_empty(row[index])]
    if not values:
        return DatasetFieldType._string

    if all(_can_be_bool(value) for value in values):
        return DatasetFieldType._bool
    if all(_can_be_int(value) for value in values):
        return DatasetFieldType._int
    if all(_can_be_float(value) for value in values):
        return DatasetFieldType._float
    return DatasetFieldType._string


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == '':
        return True
    return False


def _can_be_bool(value: Any) -> bool:
    try:
        _coerce_dataset_value(DatasetFieldType._bool, value)
        return True
    except Exception:
        return False


def _can_be_int(value: Any) -> bool:
    try:
        _coerce_dataset_value(DatasetFieldType._int, value)
        return True
    except Exception:
        return False


def _can_be_float(value: Any) -> bool:
    try:
        _coerce_dataset_value(DatasetFieldType._float, value)
        return True
    except Exception:
        return False
